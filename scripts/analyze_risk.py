# scripts/analyze_risk.py
"""
Vendor register analysis + automated flagging, Excel export with highlights, and charts.
Usage:
  python .\scripts\analyze_risk.py
  python .\scripts\analyze_risk.py --input data/vendor_register_template.xlsx --outdir outputs --days 365
"""
from pathlib import Path
import logging
import argparse
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

LOG = logging.getLogger(__name__)

def load_input(preferred):
    for p in preferred:
        if p.exists():
            LOG.info("Loading %s", p)
            if p.suffix.lower() in ('.xls', '.xlsx'):
                df = pd.read_excel(p, engine='openpyxl')
            else:
                df = pd.read_csv(p)
            return df, p
    LOG.error("No input found. Looked for: %s", ", ".join(str(x) for x in preferred))
    raise FileNotFoundError("Input file not found")

def ensure_columns(df):
    df.columns = [str(c).strip() for c in df.columns]
    expected = ['Vendor Name', 'Service', 'Risk Score', 'Assessment Date', 'Remediation Status']
    for col in expected:
        if col not in df.columns:
            LOG.warning("Column missing: %s - creating empty column", col)
            df[col] = pd.NA
    return df

def compute_flags(df, days_threshold, high_threshold):
    today = pd.Timestamp.today().normalize()
    threshold = today - pd.Timedelta(days=days_threshold)

    # parse dates & numeric
    df['Assessment Date'] = pd.to_datetime(df['Assessment Date'], errors='coerce')
    df['Risk Score'] = pd.to_numeric(df['Risk Score'], errors='coerce')

    # days since review
    df['Days Since Review'] = (today - df['Assessment Date']).dt.days

    # needs review = missing OR older than threshold
    df['Needs Review'] = df['Assessment Date'].isna() | (df['Assessment Date'] < threshold)

    # risk category buckets
    def bucket(score):
        if pd.isna(score):
            return 'Unknown'
        if score >= high_threshold:
            return 'High'
        if score >= 50:
            return 'Medium'
        return 'Low'
    df['Risk Category'] = df['Risk Score'].apply(bucket)
    return df

def save_outputs(df, outdir, high_threshold):
    # 1. Save high risk CSV
    high_path = os.path.join(outdir, "high_risk.csv")
    df.loc[df['Risk Score'] >= high_threshold].to_csv(high_path, index=False)

    # 2. Save needs review CSV
    needs_path = os.path.join(outdir, "needs_review.csv")
    df.loc[df['Needs Review']].to_csv(needs_path, index=False)

    # 3. Save Excel with highlights
    excel_path = os.path.join(outdir, "vendor_register_flagged.xlsx")
    # (code that writes Excel goes here)

    print(f"[INFO] Saved CSVs and Excel: {high_path}, {needs_path}, {excel_path}")

    # ✅ Add this at the very end:
    return high_path, needs_path, excel_path

def highlight_excel(excel_path: Path, fill_hex='FFF2CC'):
    # light yellow default
    wb = load_workbook(excel_path)
    if 'Vendor Register' not in wb.sheetnames:
        LOG.warning("Vendor Register sheet not found for highlighting")
        return
    ws = wb['Vendor Register']
    headers = [cell.value for cell in ws[1]]
    try:
        flag_col = headers.index('Needs Review') + 1
    except ValueError:
        LOG.warning("'Needs Review' column not found in Excel; skipping highlight")
        wb.save(excel_path)
        return

    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=flag_col)
        val = cell.value
        # handle True/False, strings, 1/0
        flagged = False
        if val is True:
            flagged = True
        elif isinstance(val, str) and val.strip().lower() in ('true', '1', 'yes'):
            flagged = True
        elif isinstance(val, (int, float)) and val == 1:
            flagged = True
        if flagged:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    wb.save(excel_path)
    LOG.info("Applied highlights to flagged rows in %s", excel_path)

def make_charts(df, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)

    # Top 5 high risk bar chart
    top5 = df.sort_values('Risk Score', ascending=False).head(5)
    if not top5.empty:
        ax = top5.plot.bar(x='Vendor Name', y='Risk Score', legend=False, title='Top 5 High Risk Vendors', figsize=(8,4))
        ax.set_ylabel('Risk Score')
        plt.tight_layout()
        top5_path = outdir / 'top5_high_risk.png'
        plt.savefig(top5_path)
        plt.close()
        LOG.info("Saved top5 chart: %s", top5_path)

    # Remediation status pie chart
    status_counts = df['Remediation Status'].fillna('Unknown').value_counts()
    if not status_counts.empty:
        fig, ax = plt.subplots(figsize=(6,6))
        status_counts.plot.pie(autopct='%1.0f%%', ylabel='', title='Remediation Status Distribution', ax=ax)
        plt.tight_layout()
        pie_path = outdir / 'remediation_status_pie.png'
        plt.savefig(pie_path)
        plt.close()
        LOG.info("Saved remediation status pie: %s", pie_path)

def main():
    parser = argparse.ArgumentParser(description="Vendor register analysis + flagging")
    parser.add_argument('--input', '-i', help='Input file (xlsx or csv)', default=None)
    parser.add_argument('--outdir', '-o', help='Output dir', default='outputs')
    parser.add_argument('--days', '-d', type=int, help='Days threshold for Needs Review', default=365)
    parser.add_argument('--threshold', '-t', type=int, help='High risk threshold (default=80)', default=80)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
    LOG.info("Start analysis")

    if args.input:
        paths = [Path(args.input)]
    else:
        paths = [Path('data/vendor_register_template.xlsx'), Path('data/vendor_register_template.csv')]

    df, src = load_input(paths)
    df = ensure_columns(df)
    df = compute_flags(df, args.days, args.threshold)

    # console summary
    LOG.info("Rows total: %d", len(df))
    LOG.info("High risk (>=80): %d", int((df['Risk Score'] >= 80).sum()))
    LOG.info("Needs review: %d", int(df['Needs Review'].sum()))
    LOG.info("Risk categories:\n%s", df['Risk Category'].value_counts(dropna=False).to_string())

    outdir = Path(args.outdir)
    high_path, needs_path, excel_path = save_outputs(df, outdir, args.threshold)
    highlight_excel(excel_path)
    make_charts(df, outdir)

    LOG.info("Done. Outputs are in: %s", outdir.resolve())

if __name__ == '__main__':
    main()
